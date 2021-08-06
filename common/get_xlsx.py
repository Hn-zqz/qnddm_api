#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Author ：^_^
# @Time：2020/1/16 9:08
# @Email:2560500412@qq.com

import openpyxl
def row_column_read_datas(file,sheet):
    wk=openpyxl.load_workbook(file)
    sh=wk[sheet]
    max_column=sh.max_column
    max_row=sh.max_row
    list=[]
    dict={}
    for j in range(1,max_row+1): #选择行，初始行为1
        for i in range(1,max_column+1):#选择列 初始列为1
            a=sh.cell(row=j,column=i).value #输出当行为1时的所有列，第一次循环完毕。第二次循环行为2，输出行为2时的所有列
            list.append(a)
        dict[list[0]]=(list[1])
        list=[]
    return dict

print (row_column_read_datas(path,"Sheet1"))
