#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Author ：^_^
# @Time：2020/1/6 14:41
# @Email:2560500412@qq.com

import time
import random
import openpyxl
from time import sleep
# class Excle:
import requests
from lxml import etree


def row_column_read_data(file,sheet,row,column):
    wk=openpyxl.load_workbook(file)
    sh=wk[sheet]
    print(sh.cell(row=row,column=column).value)

def row(file,sheet,row):
    wk=openpyxl.load_workbook(file)
    sh=wk[sheet]
    max_column=sh.max_column
    rows=[]
    for i in range(1,max_column+1):
        result=sh.cell(row=row,column=i).value
        rows.append(result)
    print(rows)

def column(file,sheet,column):
    wk=openpyxl.load_workbook(file)
    sh=wk[sheet]
    max_row=sh.max_row
    columns=[]
    for i in range(1,max_row+1):
        result=sh.cell(row=i,column=column).value
        columns.append(result)
    print(columns)

def row_column_write_data(file,sheet,row,column,value):
    wk=openpyxl.load_workbook(file)
    sh=wk[sheet]
    sh.cell(row=row,column=column,value=value)
    wk.save(file)

def row_column_write_datas(file,sheet,row,column,value):
    wk=openpyxl.load_workbook(file)
    sh=wk[sheet]
    max_column = sh.max_column
    max_row=sh.max_row
    for i in range(1,max_row+1):
        for j in range(1,max_column+1):
            sh.append()
    sh.cell(row=row,column=column,value=value)
    wk.save(file)

# def row_column_read_datas_login(file,sheet):
#     wk=openpyxl.load_workbook(file)
#     sh=wk[sheet]
#     max_column=sh.max_column
#     max_row=sh.max_row
#     list=[]
#     row_list=[]
#     for j in range(2,max_row+1): #选择行，初始行为1
#         for i in range(1,max_column+1):#选择列 初始列为1
#             a=sh.cell(row=j,column=i).value #输出当行为1时的所有列，第一次循环完毕。第二次循环行为2，输出行为2时的所有列
#             list.append(a)
#         yu=(list[0],list[1],list[2])
#         list=[]
#         row_list.append(yu)
#     return row_list
#
# def row_column_read_datas_sign(file,sheet):
#     wk=openpyxl.load_workbook(file)
#     sh=wk[sheet]
#     max_column=sh.max_column
#     max_row=sh.max_row
#     list=[]
#     row_list=[]
#     for j in range(2,max_row+1): #选择行，初始行为1
#         for i in range(1,max_column+1):#选择列 初始列为1
#             a=sh.cell(row=j,column=i).value #输出当行为1时的所有列，第一次循环完毕。第二次循环行为2，输出行为2时的所有列
#             list.append(a)
#         yu=(list[0],list[1],list[2],list[3])
#         list=[]
#         row_list.append(yu)
#     return row_list
#
# def row_column_read_datas_(file,sheet):
#     wk=openpyxl.load_workbook(file)
#     sh=wk[sheet]
#     max_column=sh.max_column
#     max_row=sh.max_row
#     # list=[]
#     # row_list=[]
#     for j in range(2,max_row+1): #选择行，初始行为1
#         for i in range(1,max_column+1):#选择列 初始列为1
#             a=sh.cell(row=j,column=i).value #输出当行为1时的所有列，第一次循环完毕。第二次循环行为2，输出行为2时的所有列
#     #         list.append(a)
#     #     yu=(list[0],list[1],list[2])
#     #     list=[]
#     #     row_list.append(yu)
#     # return row_list

def row_column_read_datas(file,sheet,row=1,column=1):
    wk=openpyxl.load_workbook(file)
    sh=wk[sheet]
    max_column=sh.max_column
    max_row=sh.max_row
    list=[]
    row_list=[]
    for j in range(row,max_row+1): #选择行，初始行为1
        for i in range(column,max_column+1):#选择列 初始列为1
            a=sh.cell(row=j,column=i).value #输出当行为1时的所有列，第一次循环完毕。第二次循环行为2，输出行为2时的所有列
            list.append(a)
        yu=tuple(list)
        list=[]
        row_list.append(yu)
    return row_list

def get_dict(file,sheet,row=1):
    wk = openpyxl.load_workbook(file)
    sh = wk[sheet]
    max_row = sh.max_row
    dict={}
    for i in range(row,max_row+1):
        name=sh.cell(row=i,column=1).value
        value=sh.cell(row=i,column=2).value
        dict[name]=value
    return dict


class Swip():
    def __init__(self,wk):
        self.wk=wk

    def Swipleft(self,n=1,t=1000):
        size=self.wk.get_window_size()
        s_x=int(size["width"]*0.9)
        s_y=int(size["height"]*0.5)
        e_x=int(size["width"]*0.1)
        e_y=s_y
        for i in range(n):
            self.wk.swipe(s_x,s_y,e_x,e_y,t)
            sleep(1)

    def Swipright(self,n=1,t=1000):
        size=self.wk.get_window_size()
        s_x=int(size["width"]*0.1)
        s_y=int(size["height"]*0.5)
        e_x=int(size["width"]*0.9)
        e_y=s_y
        for i in range(n):
            self.wk.swipe(s_x,s_y,e_x,e_y,t)
            sleep(1)

    def Swipup(self,n=1,t=1000):
        size=self.wk.get_window_size()
        s_x=int(size["width"]*0.5)
        s_y=int(size["height"]*0.9)
        e_x=s_x
        e_y=int(size["height"]*0.1)
        for i in range(n):
            self.wk.swipe(s_x,s_y,e_x,e_y,t)
            sleep(1)

    def Swipdown(self,n=1,t=1000):
        size=self.wk.get_window_size()
        s_x=int(size["width"]*0.5)
        s_y=int(size["height"]*0.1)
        e_x=s_x
        e_y=int(size["height"]*0.9)
        for i in range(n):
            self.wk.swipe(s_x,s_y,e_x,e_y,t)
            sleep(1)

def now_to_timestamp(digits = 10):
    time_stamp = time.time()
    digits = 10 ** (digits -10)
    time_stamp = int(round(time_stamp*digits))
    return time_stamp

def get_record(headers):
    url = "http://192.168.2.129:8080/users/editviewAction?record=&invalid_ie_cache=" + str(now_to_timestamp(digits=13))
    r = requests.get(url=url, headers=headers)
    tree = etree.HTML(r.text)
    return tree.xpath("//input[@name='record']/@value")[0]

def get_account():
    account=""
    for i in range(6):
        account=account+str(random.randint(1,9))
    return account

def get_mobile():
    mobile="176"
    for i in range(8):
        mobile=mobile+str(random.randint(1,9))
    return int(mobile)
